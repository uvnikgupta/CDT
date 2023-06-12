import multiprocessing
from functools import partial
import uuid, math, time, datetime, pickle, os, shutil, json
import numpy as np
import pandas as pd
import networkx as nx
from networkx.drawing.nx_agraph import graphviz_layout
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.drawing.image import Image
from io import BytesIO
from PIL import Image as PILImage
from sklearn.preprocessing import StandardScaler
from cdt.causality.graph import (GS, GIES, PC, SAM, IAMB, Inter_IAMB, 
                                 Fast_IAMB, MMPC, LiNGAM, CAM, CCDr)
from cdt.metrics import precision_recall, SID, SHD
from scm.dynamic_scm import DynamicSCM
from scmodels import SCM

plt.switch_backend('agg')

# DAG configurations
configs = [
    # {
    #     "name" : "config_11",
    #     "type" : 1,
    #     "nodes": [10,5,2],
    #     "dSCM" : 'DynamicSCM(min_parents=2, parent_levels_probs=[0.9, 0.3], \
    #         distribution_type=1, simple_operations={"+": 1})'
    # },
    # {
    #     "name" : "config_21",
    #     "type" : 2,
    #     "nodes": [10,5,2],
    #     "dSCM" : 'DynamicSCM(min_parents=2, parent_levels_probs=[0.9, 0.3], \
    #         distribution_type=2, simple_operations={"+": 1})'
    # },
    # {
    #     "name" : "config_31",
    #     "nodes": [10,5,2],
    #     "dSCM" : 'DynamicSCM(min_parents=2, parent_levels_probs=[0.9, 0.3], \
    #         simple_operations={"+": 1})'
    # },
    # {
    #     "name" : "config_12",
    #     "type" : 1,
    #     "nodes": [2,2,2,2,2,1],
    #     "dSCM" : 'DynamicSCM(min_parents=2, max_parents=2, parent_levels_probs=[1], \
    #         distribution_type=1, simple_operations={"+": 1})'
    # },
    # {
    #     "name" : "config_22",
    #     "nodes": [2,2,2,2,2,1],
    #     "type" : 2,
    #     "dSCM" : 'DynamicSCM(min_parents=2, max_parents=2, parent_levels_probs=[1], \
    #         distribution_type=2, simple_operations={"+": 1})'
    # },
    # {
    #     "name" : "config_32",
    #     "nodes": [2,2,2,2,2,1],
    #     "dSCM" : 'DynamicSCM(min_parents=2, max_parents=2, parent_levels_probs=[1], \
    #         simple_operations={"+": 1})'
    # },
    # {
    #     "name" : "config_23",
    #     "type" : 2,
    #     "nodes": [2,2,2,2,2,1],
    #     "dSCM" : 'DynamicSCM(min_parents=2, max_parents=2, \
    #         parent_levels_probs=[0.4,0.1,0.5], distribution_type=2)'
    # },
    # {
    #     "name" : "config_33",
    #     "nodes": [2,2,2,2,2,1],
    #     "dSCM" : 'DynamicSCM(min_parents=2, max_parents=2, \
    #         parent_levels_probs=[0.4,0.1,0.5])'
    # },
    {
        "name" : "config_13",
        "type" : 1,
        "nodes": 40,
        "dSCM" : 'DynamicSCM(distribution_type=1)'
    },
    {
        "name" : "config_24",
        "type" : 2,
        "nodes": 40,
        "dSCM" : 'DynamicSCM(distribution_type=2)'
    },
    {
        "name" : "config_34",
        "nodes": 40,
        "dSCM" : 'DynamicSCM()'
    },
    {
        "name" : "config_35",
        "nodes": [150, 50, 10, 3],
        "dSCM" : 'DynamicSCM(max_parents=20, complex_operations={})'
    }
]

# CDT algorithms
models = [
    {
        "name" : "GIES",
        "model" : "GIES(score='obs')",
        "parallel" : True,
        "type" : 1
    },
    {
        "name" : "GIES",
        "model" : "GIES(score='int')",
        "parallel" : True,
        "type" : 2
    },
    {
        "name" : "LiNGAM",
        "model" : "LiNGAM()",
        "parallel" : True,
        "type" : 2
    },    
    {
        "name" : "PC",
        "model" : "PC()",
        "parallel" : True
    },
    {
        "name" : "GS", 
        "model" : "GS()",
        "parallel" : True
    },
    {
        "name" : "IAMB",
        "model" : "IAMB()",
        "parallel" : True
    },
    {
        "name" : "Fast_IAMB",
        "model" : "Fast_IAMB()",
        "parallel" : True
    },
    {
        "name" : "Inter_IAMB",
        "model" : "Inter_IAMB()",
        "parallel" : True
    },
    {
        "name" : "MMPC",
        "model" : "MMPC()",
        "parallel" : True
    },
    # {
    #     "name" : "CCDr",
    #     "model" : "CCDr()",
    #    "parallel" : True
    # },
    # {
    #     "name" : "SAM",
    #     "model" : "SAM()",
    #     "parallel" : False,
    #     "scale_data" : True
    # }
]

# Global variables
plots_file = "logs/cdt_algo_plots.xlsx"
plots_sheet_name = "plots"
config_sheet_name = "configs"
log_file ='logs/cdt_algo_eval.log'
config_file = "configs_for_cdt_algo_eval.json"
metrics = {
    "aupr":'round(precision_recall(orig_dag, algo_dag)[0], 2)', 
    "sid":'int(SID(orig_dag, algo_dag))',
    "shd":'int(SHD(orig_dag, algo_dag))',
    "rt":'round(time.time() - start, 2)'}
sample_sizes = [1000, 10000, 20000, 50000, 100000]
num_iterations = 3
row = 1
   
def log_progress(message):
    with open(log_file, "a+") as f:
        ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        f.writelines(f"{ts} - {message}\n")

def consolidate_scores(scores):
    for algo in scores.keys():
        for metric in metrics.keys():
            mean = round(np.array(scores[algo][metric]).mean(), 2)
            std = round(np.array(scores[algo][metric]).std(), 2)
            scores[algo][metric] = (mean, std)

def update_scores(scores, results):
    for result in results:
        algo_name = result[0]
        with open(result[1], "rb") as file:
            dag = pickle.load(file)
        scores[algo_name]['dag'] = dag
        for metric in metrics:
            scores[algo_name][metric].append(result[2][metric])

def get_algo_metrics(algo_name, orig_dag, algo_dag, start):
    retval ={}
    if algo_dag is not None:
        for metric, formula in metrics.items():
            if "sam" in algo_name.lower() and "sid" in metric.lower():
                # Calculating SID for DAGs generated by SAM does not finish
                # in reasonable time. Hence setting this metric to inf for SAM
                retval[metric] = np.inf
            else:
                retval[metric] = eval(formula)
    else:
        for metric, formula in metrics.items():
            retval[metric] = 0
    return retval

def scale_data(data):
    columns = data.columns
    scaler = StandardScaler()
    data = scaler.fit_transform(data)
    data = pd.DataFrame(data=data, columns=columns)
    return data

def train_algo(algo_meta_data, data_file, conf_name, orig_scm_dists, iter):
    uid = uuid.uuid4()
    scm = SCM(orig_scm_dists)
    algo_name = algo_meta_data["name"]

    with open(data_file, "rb") as file:
        data = pickle.load(file)

    if "scale_data" in algo_meta_data and algo_meta_data["scale_data"]:
        data = scale_data(data)

    context = f"{algo_name} for {conf_name} using {len(data)} samples - {uid}"
    log_progress(f"Start Training : {context}")
    
    start = time.time()
    algo = eval(algo_meta_data["model"])
    try:
        algo_dag = algo.predict(data)
        metrics = get_algo_metrics(algo_name, scm.dag, algo_dag, start)
        log_progress(f"End Training : {context}")
    except Exception as e:
        algo_dag = f"EXCEPTION in training {context}"
        metrics = get_algo_metrics(algo_name, None, None, 0)
        log_progress(f"EXCEPTION in training {context}\n{e}")

    dag_file_name = f"temp/{algo_meta_data['name']}_{iter}.dag"
    with open(dag_file_name, "wb") as file:
        pickle.dump(algo_dag, file)

    return ((algo_meta_data['name'], dag_file_name, metrics))
    
def get_timeout_value(conf, num_samples):
    if isinstance(conf["nodes"], list):
        num_nodes = sum(conf["nodes"])
    else:
        num_nodes = conf["nodes"]
    
    timeout = int(math.exp(int(math.log(num_nodes * 2)) + 
                           int(math.log(math.sqrt((num_samples))))))
    return timeout

def init_scores(models):
    scores = {}
    for algo_meta_data in models:
        algo_name = algo_meta_data["name"]
        scores[algo_name] = {}
        scores[algo_name]["dag"] = None
        for metric in metrics:
            scores[algo_name][metric] = []
    return scores

def train_algos_for_sample_size(algos, data_file_names, num_samples, 
                                conf, orig_scm_dists):
    scores = init_scores(algos)
    num_processes = max(1, min(len(algos), multiprocessing.cpu_count() - 1))
    timeout = get_timeout_value(conf, num_samples)

    for iter in range(num_iterations):
        log_progress(f"\n**** Starting training iteration {iter} for \
{num_samples} samples ****")
        data_file = data_file_names[iter]
        args_list = [(algo_meta_data, data_file, conf["name"], orig_scm_dists, iter) 
                     for algo_meta_data in algos if algo_meta_data["parallel"] == True]
        try:
            pool = multiprocessing.Pool(processes=num_processes)
            results = [pool.apply_async(train_algo, args=args) 
                        for args in args_list]
            results = [result.get(timeout=timeout) for result in results]

            # Train the algos that cannot be run in parallel. These are the 
            # NN based algos that spawn their own process hence cannot be run
            # from within an already spawned process
            for algo_meta_data in algos:
                if not algo_meta_data["parallel"]:
                    result = train_algo(algo_meta_data, data_file, 
                                        conf["name"], orig_scm_dists, iter)
                    results.append(result)

            update_scores(scores, results)
        except multiprocessing.TimeoutError:
            log_progress(f"**TIMEOUT after {timeout} secs in \
iteration {iter} for {conf['name']} using {num_samples}")
            
        log_progress(f"\n==== Completed training iteration {iter} for \
{num_samples} samples ====")
    return scores

def generate_data(num_samples, scm_dists, conf_name, iter):
    scm = SCM(scm_dists)
    trials = 10
    while trials:
        msg = f"{num_samples} samples for {conf_name}"
        try:
            log_msg = f"Iter {iter}: Starting data generation of {msg}"
            log_progress(log_msg)
            data = scm.sample(num_samples)
            # data = scale_data(data)

            file_name = f"temp/data_{iter}.pkl"
            with open(file_name, "wb") as file:
                pickle.dump(data, file)

            log_msg = f"Iter {iter}: Completed data generation of {msg}"
            log_progress(log_msg)
            break
        except Exception as e:
            log_msg = f"Iter {iter}: EXCEPTION in data generation. Trying again {msg}"
            log_progress(log_msg)
            trials -= 1
            continue
    return file_name

def get_algos_for_training(conf):
    algos = []
    for algo_meta_data in models:
        if "type" in algo_meta_data:
            if "type" in conf and algo_meta_data["type"] == conf["type"]:
                # add type specifc algos
                algos.append(algo_meta_data)
            elif "type" not in conf and algo_meta_data["type"] == 1:
                # if conf is of mixed type add continuous alogs
                algos.append(algo_meta_data)
        else: # add common algos
            algos.append(algo_meta_data)
    return algos

def populate_cdt_algos_scores_for_config(scm_dists, conf):
    global row
    algos = get_algos_for_training(conf)

    iters = [(i,) for i in range(num_iterations)]
    num_processes = max(1, min(num_iterations, multiprocessing.cpu_count() - 1))

    # Run algo training for various sample sizes for the config
    for num_samples in sample_sizes:
        # Run parallel data generation for all iterations of algo training
        # Cannot generate the data while training the algos in parallel as
        # each algorithm will then have a different set of data!
        partial_worker = partial(generate_data, num_samples, scm_dists, conf["name"])
        pool = multiprocessing.Pool(processes=num_processes)
        file_names = pool.starmap(partial_worker, iters)
        pool.close()
        pool.join()

        # Train CDT algos and get their DAGs and run durations
        scores = train_algos_for_sample_size(algos, file_names, 
                                             num_samples, conf, scm_dists)
        consolidate_scores(scores)
        save_plots_and_scores(plots_file, 1, scores, conf, num_samples)  
        row = row + 1
        
def save_config_to_xl(configs_sheet, conf):
    if configs_sheet.max_row == 1:
        row = 1
    else:
        row = configs_sheet.max_row + 1
    configs_sheet.cell(row=row, column=1, value=str(conf))    

def get_score_text(score_data):
    text = ""
    for metric, values in score_data.items():
        if metric != "dag":
            text += f"{metric}:{values}\n"
    return text

def get_node_colors(dag):
    nodes = dag.nodes()
    colors = ['#F9E79F','#A3E4D7','#D7BDE2','#A9CCE3','#DAF7A6',
              '#EBDEF0','#EDBB99','#3498DB','#EC7063','#E5E8E8',]
    node_colors = []
    prev_level = ""
    index = -1
    for node in nodes:
        cur_level = node[0]
        if cur_level != prev_level:
            prev_level = cur_level
            index += 1
        if index > len(colors) - 1:
            index = 0
        node_colors.append(colors[index])
    return node_colors

def save_plots_and_scores_to_xl(plots_sheet, col, scores, conf, num_samples=""):
    for algo, score_data in scores.items():
        # Since all the DAGs have the same number of nodes with the 
        # same names, node colors will of any one DAG will be applicable
        # for all DAGs. Since the only way to access an item in a python 
        # dict is via the key, hence this sing loop for loop
        node_colors = get_node_colors(score_data["dag"])
        break

    for algo, score_data in scores.items():
        dag = score_data["dag"]
        fig, ax = plt.subplots(1, 1, figsize=(3,2), dpi=150)
        ax.set_title(f"{conf['name']}_{algo}_{num_samples}")

        if isinstance(dag, nx.classes.digraph.DiGraph):
            pos = graphviz_layout(dag, prog="dot")
            nx.draw(dag, pos=pos, ax=ax, with_labels=True, node_size=150, 
                    node_color=node_colors, font_size=7, alpha=0.5)
            text = get_score_text(score_data)
            fig.text(0.65, 0.05, text, fontsize=10, color='black')
        else:
            text = dag
            fig.text(0.2, 0.2, text, fontsize=8, color='red')
        
        fig.savefig("plot.png")
        plt.close(fig)
        
        pil_image = PILImage.open('plot.png')
        image_bytes = BytesIO()
        pil_image.save(image_bytes, format='PNG')
        image_bytes.seek(0)
        image = Image(image_bytes)

        target_cell = chr(65 + col) + str(row)
        width, height = 35, 110
        plots_sheet.column_dimensions[target_cell[0]].width = width
        plots_sheet.row_dimensions[int(target_cell[1:])].height = height
        image.width = width * 6.8
        image.height = height * 1.3

        plots_sheet.add_image(image, target_cell)
        pil_image.close()
        os.remove('plot.png')
        col = col + 1

def check_num_rows_in_xl_and_move_it(plots_file):
    global row
    if row > 10:
        prefix = datetime.datetime.now().strftime("%d%H%M%S")
        new_file_path = f"{os.path.splitext(plots_file)[0]}_{prefix}.xlsx"
        shutil.move(plots_file, new_file_path)
        row = 1
        log_progress(f"The file '{plots_file}' has been renamed to '{new_file_path}' \
due to exceeding the maximum size.")

def get_xl_workbook_for_plots(plots_file):
    workbook = None
    if os.path.exists(plots_file):
        check_num_rows_in_xl_and_move_it(plots_file)

    try:
        workbook = openpyxl.load_workbook(plots_file)
    except:
        workbook = openpyxl.Workbook()
        plots_sheet = workbook.active
        plots_sheet.title = plots_sheet_name
        configs_sheet = workbook.create_sheet()
        configs_sheet.title = config_sheet_name
    return workbook

def save_plots_and_scores(plots_file, col, scores, conf, num_samples=""):
    log_progress("Saving DAGs - Start")
    workbook = get_xl_workbook_for_plots(plots_file)
    
    plots_sheet = workbook[plots_sheet_name]
    save_plots_and_scores_to_xl(plots_sheet, col, scores, conf, num_samples)

    if col == 0:
        configs_sheet = workbook[config_sheet_name]
        save_config_to_xl(configs_sheet, conf)

    workbook.save(plots_file)
    workbook.close()
    log_progress("Saving DAGs - Done")

def get_scm(config):
    input_nodes = config["nodes"]
    dSCM = eval(config["dSCM"])
    scm = dSCM.create(input_nodes)
    return scm, dSCM.get_scm_dists()

def get_algo_eval_configs(config_file):
    with open(config_file,"r") as file:
        configs = json.load(file)
    return configs
    
if __name__ == "__main__":
    temp_folder = "temp"
    if not os.path.exists(temp_folder):
        os.makedirs(temp_folder)

    for filename in os.listdir("logs"):
        file_path = os.path.join("logs", filename)
        if os.path.isfile(file_path):
            os.remove(file_path)

    for conf in get_algo_eval_configs(config_file):
        scm, scm_dists = get_scm(conf)
        save_plots_and_scores(plots_file, 0, 
                           {"Original":{"dag": scm.dag}}, 
                           conf)
        
        populate_cdt_algos_scores_for_config(scm_dists, conf)