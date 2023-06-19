import os, yaml, pickle
import pandas as pd
import numpy as np
import openpyxl
import networkx as nx
from networkx.drawing.nx_agraph import graphviz_layout
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
from PIL import Image as PILImage
from scmodels import SCM

# Workaround for pygraphviz where it crashes after created 170 graphs
# ref : https://stackoverflow.com/questions/60876623/pygraphviz-crashes-after-drawing-170-graphs
import win32file as wfile
wfile._setmaxstdio(1024)

# fix for tkinter error while creating plots
plt.switch_backend('agg')

class ReportGenerator():
    def __init__(self, status_file, report_file, orig_dists_dir, config_file) -> None:
        self.__status_file = status_file
        self.__report_file = report_file
        self.__orig_dists_dir = orig_dists_dir
        self.__config_file = config_file
        self.__plots_sheet_name = "plots"
        self.__config_sheet_name = "configs"
        self.__details_sheet_name = "details"

    def __save_config_to_xl(self):
        with open(self.__config_file) as file:
            configs = yaml.safe_load(file)

        workbook = self.__get_xl_workbook()
        sheet = workbook[self.__config_sheet_name]

        for i, config in enumerate(configs):
            target_cell = "A" + str(i + 1)
            sheet[target_cell] = str(config)

        workbook.save(self.__report_file)
        workbook.close()

    def __write_tabular_report_to_xl(self, df):
        workbook = self.__get_xl_workbook()
        sheet = workbook[self.__details_sheet_name]
        for row in dataframe_to_rows(df, index=False, header=True):
            sheet.append(row)
        workbook.save(self.__report_file)
        workbook.close()

    def __get_node_colors(self, dag):
        nodes = dag.nodes()
        colors = ['#F9E79F','#A3E4D7','#D7BDE2','#A9CCE3','#DAF7A6',
                  '#EBDEF0','#EDBB99','#3498DB','#EC7063','#E5E8E8',]
        node_colors = []
        prev_level = ""
        index = -1
        for node in nodes:
            cur_level = node[0]
            if cur_level != prev_level:
                prev_level = cur_level
                index += 1
            if index > len(colors) - 1:
                index = 0
            node_colors.append(colors[index])
        return node_colors

    def __get_score_text(self, plot_data):
        text = ""
        if "aupr" in plot_data.keys():
            text = f'{plot_data["aupr"]}\n{plot_data["sid"]}\n{plot_data["shd"]}\n{plot_data["rt"]}'
        return text
    
    def __create_graphs(self, sheet, plot_data):
        title = plot_data["id"]
        dag = plot_data["dag"]
        fig, ax = plt.subplots(1, 1, figsize=(3,2), dpi=150)
        ax.set_title(title)

        if isinstance(dag, nx.classes.digraph.DiGraph):
            node_colors = self.__get_node_colors(plot_data["dag"])
            pos = graphviz_layout(dag, prog="dot")
            nx.draw(dag, pos=pos, ax=ax, with_labels=True, node_size=150, 
                    node_color=node_colors, font_size=7, alpha=0.5)
            text = self.__get_score_text(plot_data)
            fig.text(0.65, 0.05, text, fontsize=10, color='black')
        else:
            text = dag
            fig.text(0.2, 0.2, text, fontsize=8, color='red')
        
        fig.savefig("plot.png")
        plt.close(fig)
        
        pil_image = PILImage.open('plot.png')
        image_bytes = BytesIO()
        pil_image.save(image_bytes, format='PNG')
        image_bytes.seek(0)
        image = Image(image_bytes)

        target_cell = chr(65 + plot_data["col"]) + str(plot_data["row"])
        sheet[target_cell] = title
        width, height = 35, 110
        sheet.column_dimensions[target_cell[0]].width = width
        sheet.row_dimensions[int(target_cell[1:])].height = height
        image.width = width * 6.8
        image.height = height * 1.3

        sheet.add_image(image, target_cell)
        pil_image.close()
        os.remove('plot.png')

    def __get_xl_workbook(self):
        workbook = None
        try:
            workbook = openpyxl.load_workbook(self.__report_file)
        except:
            workbook = openpyxl.Workbook()
            plots_sheet = workbook.active
            plots_sheet.title = self.__plots_sheet_name
            configs_sheet = workbook.create_sheet()
            configs_sheet.title = self.__config_sheet_name
            details_sheet = workbook.create_sheet()
            details_sheet.title = self.__details_sheet_name
        return workbook

    def __save_plots_to_xl(self, data):
        workbook = self.__get_xl_workbook()

        for _, row_data in data.iterrows():
            plot_data = {}
            plots_sheet = workbook[self.__plots_sheet_name]
            plot_data["row"] = row_data["row"]

            if row_data["config_col"] == 0:
                plot_data["col"] = 0
                plot_data["id"] = row_data["config"]
                dists_file = os.path.join(self.__orig_dists_dir, row_data["orig_dist"])
                with open(dists_file, "rb") as file:
                    dists = pickle.load(file)
                    scm = SCM(dists)
                plot_data['dag'] = scm.dag
                self.__create_graphs(plots_sheet, plot_data)

            plot_data["col"] = row_data["col"]
            plot_data["id"] = row_data["id"]
            plot_data["aupr"] = f'aupr: {row_data["aupr_mean"]}, {row_data["aupr_std"]}'
            plot_data["sid"] = f' sid: {row_data["sid_mean"]}, {row_data["sid_std"]}'
            plot_data["shd"] = f' shd: {row_data["shd_mean"]}, {row_data["shd_std"]}'
            plot_data["rt"] = f'  rt: {row_data["rt_mean"]}, {row_data["rt_std"]}'
            
            with open(row_data["dag"], "rb") as file:
                algo_dag = pickle.load(file)
            plot_data["dag"] = algo_dag

            self.__create_graphs(plots_sheet, plot_data)
            print(plot_data["row"])

        workbook.save(self.__report_file)
        workbook.close()

    def __get_status_df(self):
        with open(self.__status_file, "r") as file:
            status = yaml.safe_load(file)
        
        status_df = pd.DataFrame(status).T.reset_index()
        status_df["id"] = (status_df["index"]
                                .str.split("_")
                                .str[:-1]
                                .str.join("_"))
        status_df["algo"] = (status_df["id"]
                                  .str.split("_")
                                  .str[:-3]
                                  .str.join("_"))
        status_df["config"] = (status_df["id"]
                                  .str.split("_")
                                  .str[-3:-1]
                                  .str.join("_"))
        status_df["samples"] = (status_df["id"]
                                  .str.split("_")
                                  .str[-1]).astype(int)
        status_df["orig_dist"] = status_df["config"] + ".dists"
        
        status_df.drop(columns=["index"], inplace=True)
        return status_df
    
    def __get_config_nodes(self):
        config_node = {}
        with open(self.__config_file, "r") as file:
            configs = yaml.safe_load(file)

        for conf in configs:
            config_node[conf["name"]] = sum(np.ravel(np.array([conf["nodes"]])))
        return config_node
    
    def __get_status_aggregates(self):
        df = self.__get_status_df()
        gb = df.groupby("id")
        counts = gb.size().to_frame(name="counts")
        counts = (counts
                  .join(gb.agg({"aupr": "mean"}).astype(float).round(2).rename(columns={"aupr": "aupr_mean"}))
                  .join(gb.agg({"aupr": "std"}).astype(float).round(2).rename(columns={"aupr": "aupr_std"}))
                  .join(gb.agg({"sid": "mean"}).astype(float).round(2).rename(columns={"sid": "sid_mean"}))
                  .join(gb.agg({"sid": "std"}).astype(float).round(2).rename(columns={"sid": "sid_std"}))
                  .join(gb.agg({"shd": "mean"}).astype(float).round(2).rename(columns={"shd": "shd_mean"}))
                  .join(gb.agg({"shd": "std"}).astype(float).round(2).rename(columns={"shd": "shd_std"}))
                  .join(gb.agg({"rt": "mean"}).astype(float).round(2).rename(columns={"rt": "rt_mean"}))
                  .join(gb.agg({"rt": "std"}).astype(float).round(2).rename(columns={"rt": "rt_std"}))
                  .join(gb.agg({"dag": "last"}))
                  .join(gb.agg({"orig_dist": "last"}))
                  .join(gb.agg({"algo": "last"}))
                  .join(gb.agg({"config": "last"}))
                  .join(gb.agg({"samples": "last"}))
                  .reset_index())
        
        for conf, nodes in self.__get_config_nodes().items():
            counts.loc[counts["config"] == conf, "nodes"] = nodes
        counts["nodes"] = counts["nodes"].astype(int)

        counts.sort_values(by=["nodes", "config", "samples", "aupr_mean"],
                           ascending=[True, True, True, False], inplace=True)
        
        # Calculate row numbers for the plots in xl
        counts['row'] = (counts['config'].ne(counts['config'].shift()) | 
                         counts['samples'].ne(counts['samples'].shift())).cumsum()
        counts["col"] = counts.groupby(["config", "samples"]).cumcount() + 1

        # Set the column number of the plot of the original configuration to 0. 
        # This is for the first row of every new configuraion
        counts["config_col"] = ''
        counts.loc[counts['config'].ne(counts['config'].shift()), 'config_col'] = 0
        return counts
    
    def generate_report(self):
        agg = self.__get_status_aggregates()
        self.__save_plots_to_xl(agg)
        self.__save_config_to_xl()
        self.__write_tabular_report_to_xl(agg[["samples", "config", "algo", "nodes", 
                                               "aupr_mean", "aupr_std", "sid_mean", "sid_std", 
                                               "shd_mean", "shd_std", "rt_mean", "rt_std"]])
        
if __name__ == "__main__":
    config_file = "dag_generation_configs.yml"
    status_file = "logs/training_status.yml"
    report_file = "logs/cdt_algo_metrics.xlsx"
    orig_dists_dir = "data_cdt_algo_eval"
    report_gen = ReportGenerator(status_file, report_file, orig_dists_dir, config_file)
    report_gen.generate_report()